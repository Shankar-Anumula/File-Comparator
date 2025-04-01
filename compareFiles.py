import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def compare_excel_files(source_dir, target_dir, lookup_dir, output_dir="comparison_reports"):
    os.makedirs(output_dir, exist_ok=True)
    
    source_files = [f for f in os.listdir(source_dir) if f.endswith(('.xlsx', '.xls'))]
    target_files = [f for f in os.listdir(target_dir) if f.endswith(('.xlsx', '.xls'))]
    lookup_files = [f for f in os.listdir(lookup_dir) if f.endswith(('.xlsx', '.xls'))]
    
    if not source_files or not target_files or not lookup_files:
        print("No Excel files found in source, target, or lookup directories.")
        return
    
    for source_file in source_files:
        source_path = os.path.join(source_dir, source_file)
        source_name = os.path.splitext(source_file)[0]
        
        matching_target = [f for f in target_files if os.path.splitext(f)[0] == source_name]
        matching_lookup = [f for f in lookup_files if os.path.splitext(f)[0] == source_name]
        
        if not matching_lookup:
            print(f"No lookup file found for {source_file}. Skipping.")
            continue
        
        lookup_path = os.path.join(lookup_dir, matching_lookup[0])
        
        # Read both sheets from the lookup file
        source_schema = pd.read_excel(lookup_path, sheet_name='Source_Schema')
        target_schema = pd.read_excel(lookup_path, sheet_name='Target_Schema')
        
        source_columns = source_schema['Column_Name'].tolist()
        source_types = dict(zip(source_schema['Column_Name'], source_schema['Data_Type']))
        source_pks = source_schema[source_schema['Primary_Key'] == 'Yes']['Column_Name'].tolist()
        
        target_columns = target_schema['Column_Name'].tolist()
        target_types = dict(zip(target_schema['Column_Name'], target_schema['Data_Type']))
        target_pks = target_schema[target_schema['Primary_Key'] == 'Yes']['Column_Name'].tolist()
        
        if matching_target:
            target_file = matching_target[0]
            target_path = os.path.join(target_dir, target_file)
            output_file = os.path.join(output_dir, f"report_{source_name}.xlsx")
            process_comparison(source_path, target_path, source_columns, source_types, source_pks, 
                             target_columns, target_types, target_pks, output_file)
        else:
            for target_file in target_files:
                target_path = os.path.join(target_dir, target_file)
                target_name = os.path.splitext(target_file)[0]
                output_file = os.path.join(output_dir, f"report_{source_name}_vs_{target_name}.xlsx")
                process_comparison(source_path, target_path, source_columns, source_types, source_pks, 
                                 target_columns, target_types, target_pks, output_file)

def process_comparison(source_path, target_path, source_columns, source_types, source_pks, 
                      target_columns, target_types, target_pks, output_file):
    try:
        # Read the Excel files
        df1 = pd.read_excel(source_path)
        df2 = pd.read_excel(target_path)
        
        # Remove completely empty rows and columns
        df1 = df1.dropna(how='all').dropna(how='all', axis=1)
        df2 = df2.dropna(how='all').dropna(how='all', axis=1)
        
        # Handle nulls
        df1 = df1.fillna('NULL')
        df2 = df2.fillna('NULL')
        
        # Check column differences
        missing_cols1 = set(source_columns) - set(df1.columns)
        missing_cols2 = set(target_columns) - set(df2.columns)
        extra_cols1 = set(df1.columns) - set(source_columns)
        extra_cols2 = set(df2.columns) - set(target_columns)
        
        # Filter to expected columns and enforce order
        df1 = df1[source_columns] if all(col in df1.columns for col in source_columns) else df1
        df2 = df2[target_columns] if all(col in df2.columns for col in target_columns) else df2
        
        # Convert to specified data types
        for col, dtype in source_types.items():
            try:
                if col in df1.columns:
                    df1[col] = df1[col].astype(dtype)
            except Exception as e:
                print(f"Error converting {col} to {dtype} in source for {output_file}: {str(e)}")
        
        for col, dtype in target_types.items():
            try:
                if col in df2.columns:
                    df2[col] = df2[col].astype(dtype)
            except Exception as e:
                print(f"Error converting {col} to {dtype} in target for {output_file}: {str(e)}")
        
        # Remove duplicates based on primary keys if they exist
        if source_pks:
            df1 = df1.drop_duplicates(subset=source_pks).reset_index(drop=True)
        else:
            df1 = df1.drop_duplicates().reset_index(drop=True)
        
        if target_pks:
            df2 = df2.drop_duplicates(subset=target_pks).reset_index(drop=True)
        else:
            df2 = df2.drop_duplicates().reset_index(drop=True)
        
        # Determine common primary keys for merging
        common_pks = list(set(source_pks) & set(target_pks))
        if not common_pks and (source_pks or target_pks):
            print(f"Warning: No common primary keys between source {source_pks} and target {target_pks} for {output_file}. Falling back to full comparison.")
        
        # Merge based on common primary keys if they exist, otherwise full comparison
        if common_pks:
            merged = df1.merge(df2, how='outer', on=common_pks, suffixes=('_source', '_target'), indicator=True)
        else:
            merged = df1.merge(df2, how='outer', indicator=True, suffixes=('_source', '_target'))
        
        matched = merged[merged['_merge'] == 'both'].drop(columns=['_merge'])
        mismatched = merged[merged['_merge'] != 'both']
        
        # Separate missing records
        missing_in_df2 = mismatched[mismatched['_merge'] == 'left_only'].drop(columns=['_merge'])
        missing_in_df1 = mismatched[mismatched['_merge'] == 'right_only'].drop(columns=['_merge'])
        
        # Prepare mismatched with differences highlighted
        mismatched_highlight = mismatched.copy()
        all_columns = set(source_columns + target_columns) - set(common_pks)  # Exclude primary keys from highlighting
        for col in all_columns:
            source_col = f"{col}_source"
            target_col = f"{col}_target"
            if source_col in mismatched and target_col in mismatched:
                mismatched_highlight[source_col] = mismatched.apply(
                    lambda row: f"{row[source_col]}" if pd.isna(row[target_col]) or row[source_col] != row[target_col] else row[source_col], axis=1)
                mismatched_highlight[target_col] = mismatched.apply(
                    lambda row: f"{row[target_col]}" if pd.isna(row[source_col]) or row[source_col] != row[target_col] else row[target_col], axis=1)
        
        # Prepare summary
        summary = pd.DataFrame({
            'Metric': [
                'Source Rows (after deduplication)', 
                'Target Rows (after deduplication)',
                'Source Columns', 
                'Target Columns',
                'Matched Records',
                'Mismatched Records',
                'Missing in Target',
                'Missing in Source'
            ],
            'Value': [
                len(df1),
                len(df2),
                len(df1.columns),
                len(df2.columns),
                len(matched),
                len(mismatched),
                len(missing_in_df2),
                len(missing_in_df1)
            ]
        })
        
        if missing_cols1:
            summary = pd.concat([summary, pd.DataFrame({'Metric': ['Source Missing Columns'], 'Value': [str(missing_cols1)]})], ignore_index=True)
        if missing_cols2:
            summary = pd.concat([summary, pd.DataFrame({'Metric': ['Target Missing Columns'], 'Value': [str(missing_cols2)]})], ignore_index=True)
        if extra_cols1:
            summary = pd.concat([summary, pd.DataFrame({'Metric': ['Source Extra Columns'], 'Value': [str(extra_cols1)]})], ignore_index=True)
        if extra_cols2:
            summary = pd.concat([summary, pd.DataFrame({'Metric': ['Target Extra Columns'], 'Value': [str(extra_cols2)]})], ignore_index=True)
        
        # Write to Excel
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            summary.to_excel(writer, sheet_name='Summary', index=False)
            matched.to_excel(writer, sheet_name='Matched_Records', index=False)
            mismatched_highlight.to_excel(writer, sheet_name='Mismatched_Records', index=False)
            missing_in_df2.to_excel(writer, sheet_name='Missing_in_Target', index=False)
            missing_in_df1.to_excel(writer, sheet_name='Missing_in_Source', index=False)
        
        # Highlight differences in Mismatched_Records
        wb = load_workbook(output_file)
        ws = wb['Mismatched_Records']
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                col_letter = chr(64 + col)
                header = ws[f"{col_letter}1"].value
                
                if '_source' in header and not any(pk in header for pk in common_pks):
                    target_header = header.replace('_source', '_target')
                    if target_header in ws[1]:
                        target_col_idx = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)].index(target_header) + 1
                        target_cell = ws.cell(row=row, column=target_col_idx)
                        if pd.isna(target_cell.value) or (cell.value != target_cell.value and not pd.isna(cell.value)):
                            cell.fill = red_fill
                elif '_target' in header and not any(pk in header for pk in common_pks):
                    source_header = header.replace('_target', '_source')
                    if source_header in ws[1]:
                        source_col_idx = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)].index(source_header) + 1
                        source_cell = ws.cell(row=row, column=source_col_idx)
                        if pd.isna(source_cell.value) or (cell.value != source_cell.value and not pd.isna(cell.value)):
                            cell.fill = red_fill
        
        wb.save(output_file)
        print(f"Comparison complete for {os.path.basename(source_path)} vs {os.path.basename(target_path)}. Report written to {output_file}")
    
    except Exception as e:
        print(f"Error processing {os.path.basename(source_path)} vs {os.path.basename(target_path)}: {str(e)}")

# Example usage
try:
    source_dir = "path/to/source/directory"
    target_dir = "path/to/target/directory"
    lookup_dir = "path/to/lookup/directory"
    compare_excel_files(source_dir, target_dir, lookup_dir, "comparison_reports")
except Exception as e:
    print(f"An error occurred: {str(e)}")
